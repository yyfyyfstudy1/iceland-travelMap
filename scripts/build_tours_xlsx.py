#!/usr/bin/env python3
"""Build an Excel of ALL troll.is daily-tour departure (starting) prices — 12 tours."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = "/Users/mt/travelMengting/troll_daily_tours_prices.xlsx"

# (No, Name, Price USD, Unit, Hours, Rating, Type, Note, URL)
# Rows 1-10 = standard per-person small-group tours (sorted by price asc)
# Rows 11-12 = private/VIP "Spring Sale" tours (priced separately)
rows = [
    (1,  "Golden Circle, Bruarfoss & Kerid Volcanic Crater - Small Group Day Tour", 106,    "每人", 8,  4.8, "Troll.is Original", "",                          "https://troll.is/day-tour/golden-circle-bruarfoss-kerid/"),
    (2,  "Snæfellsnes Peninsula Small Group Day Tour from Reykjavík",               140,    "每人", 12, 4.9, "Troll.is Original", "",                          "https://troll.is/day-tour/snaefellsnes-peninsula/"),
    (3,  "South Coast & Glacier Hike Minibus Tour from Reykjavík",                  179,    "每人", 12, 5.0, "Troll.is Original", "",                          "https://troll.is/day-tour/south-coast-glacier-hike/"),
    (4,  "斯奈山半岛一日游 (Snæfellsnes Peninsula Chinese Day Tour)",                 179,    "每人", 12, 4.9, "Troll.is Original", "中文向导",                   "https://troll.is/day-tour/snaefellsnes-peninsula-chinese-day-tour/"),
    (5,  "Snorkeling in Silfra with Transfer from Reykjavik",                       223,    "每人", 6,  5.0, "—",                 "",                          "https://troll.is/day-tour/snorkeling-in-silfra-with-transfer-from-reykjavik/"),
    (6,  "Riding and Hiking in the Valley Reykjadalur",                             225.44, "每人", 9,  4.5, "Trusted Partner",   "",                          "https://troll.is/day-tour/riding-and-hiking-in-the-valley-reykjadalur/"),
    (7,  "Golden Circle & Blue Lagoon Day Tour from Reykjavik",                     253.8,  "每人", 11, 4.8, "Trusted Partner",   "",                          "https://troll.is/day-tour/golden-circle-blue-lagoon/"),
    (8,  "Golden Circle & Snorkel in Silfra Day Tour from Reykjavik",               289,    "每人", 10, 5.0, "Troll.is Original", "",                          "https://troll.is/day-tour/golden-circle-snorkeling-in-silfra/"),
    (9,  "South Coast & Katla Ice Cave Small Group Day Tour from Reykjavik",        299,    "每人", 12, 5.0, "Troll.is Original", "",                          "https://troll.is/day-tour/south-coast-katla/"),
    (10, "Landmannalaugar Super Jeep Day Tour from Reykjavík",                      350,    "每人", 13, 4.9, "Troll.is Original", "仅夏季 Summer",              "https://troll.is/day-tour/landmannalaugar-super-jeep-day-tour-from-reykjavik/"),
    (11, "Private VIP Golden Circle Day Tour (8h,可延长至12h)",                     1350,   "每人", 8,  4.7, "—",                 "私人VIP · Spring Sale 特价",  "https://troll.is/day-tour/vip-private-golden-circle-day-tour/"),
    (12, "VIP Private South Coast Tour (Optional Glacier Hike / Katla Ice Cave)",   1800,   "每团", 12, 5.0, "Troll.is Original", "私人VIP · 特价 · 整团报价",    "https://troll.is/day-tour/south-coast-optional-glacier-hike-katla-private-tour/"),
]

headers = ["序号", "团名 (Tour Name)", "发团起价 (USD)", "计价单位", "时长(小时)", "评分", "产品类型", "备注", "页面链接"]

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Troll.is 一日游"

FONT = "Arial"
ncols = len(headers)
last_col = get_column_letter(ncols)

navy, teal, band = "1F5C73", "2E7D91", "EAF1F3"
gold = "FFF3D6"  # highlight for private/VIP rows
thin = Side(style="thin", color="C9D6DA")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

# --- Title ---
ws.merge_cells(f"A1:{last_col}1")
t = ws["A1"]
t.value = "Troll.is 一日游 (Day Tours) — 全部 12 个团 · 发团起价一览"
t.font = Font(name=FONT, size=14, bold=True, color="FFFFFF")
t.fill = PatternFill("solid", fgColor=navy)
t.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 30

# --- Note ---
ws.merge_cells(f"A2:{last_col}2")
n = ws["A2"]
n.value = ("币种:USD 美元(网站默认显示)  |  数据抓取:2026-07-25  |  来源:troll.is/day-tour/ 归档页(经 sitemap 核对,共 12 个)  |  "
           "注:标准小团按「每人」报价;第 11–12 为私人 VIP 特价团,单独计价(其中 #12 为整团 /group 报价,不可与每人价直接比较)。价格随汇率浮动,可切换 ISK/CNY/EUR/GBP/CAD。")
n.font = Font(name=FONT, size=9, italic=True, color="52514E")
n.fill = PatternFill("solid", fgColor=band)
n.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
ws.row_dimensions[2].height = 42

# --- Header row (row 3) ---
HDR = 3
for c, h in enumerate(headers, start=1):
    cell = ws.cell(row=HDR, column=c, value=h)
    cell.font = Font(name=FONT, size=11, bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor=teal)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border
ws.row_dimensions[HDR].height = 26

# --- Data rows ---
first_data = HDR + 1
for i, r in enumerate(rows):
    rr = first_data + i
    no, name, price, unit, hours, rating, typ, note, url = r
    is_private = no >= 11
    shade = gold if is_private else ("FFFFFF" if i % 2 == 0 else band)
    for c, v in enumerate([no, name, price, unit, hours, rating, typ, note, url], start=1):
        cell = ws.cell(row=rr, column=c, value=v)
        cell.font = Font(name=FONT, size=10, color="0B0B0B")
        cell.fill = PatternFill("solid", fgColor=shade)
        cell.border = border
        cell.alignment = Alignment(vertical="center")
    ws.cell(row=rr, column=1).alignment = Alignment(horizontal="center", vertical="center")
    pc = ws.cell(row=rr, column=3)
    pc.number_format = '"$"#,##0.00'
    pc.alignment = Alignment(horizontal="right", vertical="center")
    pc.font = Font(name=FONT, size=10, bold=True, color="B23A00")
    ws.cell(row=rr, column=4).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=rr, column=5).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=rr, column=5).number_format = "0"
    ws.cell(row=rr, column=6).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=rr, column=6).number_format = "0.0"
    ws.cell(row=rr, column=7).alignment = Alignment(horizontal="center", vertical="center")
    lk = ws.cell(row=rr, column=9)
    lk.hyperlink = url
    lk.font = Font(name=FONT, size=9, color="1155CC", underline="single")

last_data = first_data + len(rows) - 1

# --- Summary block (standard per-person tours only = rows 1-10) ---
std_prices = [r[2] for r in rows if r[0] <= 10]
srow = last_data + 2
summ = [
    ("团数合计 (全部)",              len(rows),                              "0"),
    ("标准小团数 (每人计价)",         len(std_prices),                        "0"),
    ("标准小团 · 最低发团价 (USD)",   min(std_prices),                        '"$"#,##0.00'),
    ("标准小团 · 最高发团价 (USD)",   max(std_prices),                        '"$"#,##0.00'),
    ("标准小团 · 平均发团价 (USD)",   round(sum(std_prices)/len(std_prices), 2), '"$"#,##0.00'),
]
for i, (label, value, fmt) in enumerate(summ):
    r = srow + i
    lcell = ws.cell(row=r, column=2, value=label)
    lcell.font = Font(name=FONT, size=10, bold=True, color="1F5C73")
    lcell.alignment = Alignment(horizontal="right", vertical="center")
    vcell = ws.cell(row=r, column=3, value=value)
    vcell.font = Font(name=FONT, size=10, bold=True, color="0B0B0B")
    vcell.alignment = Alignment(horizontal="right", vertical="center")
    vcell.number_format = fmt
ws.cell(row=srow + len(summ), column=2,
        value="(均价不含第 11–12 私人 VIP 团)").font = Font(name=FONT, size=8, italic=True, color="898781")

# --- Column widths ---
widths = [6, 56, 15, 10, 10, 7, 17, 22, 44]
for c, w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(c)].width = w

ws.freeze_panes = f"A{first_data}"
wb.save(OUT)
print("saved", OUT, "with", len(rows), "tours")
